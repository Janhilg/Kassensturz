from pathlib import Path
from openpyxl import Workbook, load_workbook

EXCEL_HEADERS = [
    "ID",
    "Date",
    "Timestamp",
    "Event name",
    "Counted by",
    "Cash sum",
    "Event status",
    "Comment",
    "100 €",
    "50 €",
    "20 €",
    "10 €",
    "5 €",
    "2 €",
    "1 €",
    "0.50 €",
    "0.20 €",
    "0.10 €",
]


def export_entries_to_excel(entries: list[dict], excel_path: Path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Kassensturz"
    sheet.append(EXCEL_HEADERS)

    for entry in entries:
        sheet.append([
            entry.get("id", ""),
            entry.get("date", ""),
            entry.get("timestamp", ""),
            entry.get("event_name", ""),
            entry.get("counted_by", ""),
            entry.get("cash_sum", ""),
            entry.get("event_status", ""),
            entry.get("comment", ""),
            entry.get("denom_100", ""),
            entry.get("denom_50", ""),
            entry.get("denom_20", ""),
            entry.get("denom_10", ""),
            entry.get("denom_5", ""),
            entry.get("denom_2", ""),
            entry.get("denom_1", ""),
            entry.get("denom_050", ""),
            entry.get("denom_020", ""),
            entry.get("denom_010", ""),
        ])

    excel_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(excel_path)


def import_entries_from_excel(excel_path: Path) -> list[dict]:
    workbook = load_workbook(excel_path, data_only=True)
    sheet = workbook.active

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    header = [str(cell) if cell is not None else "" for cell in rows[0]]
    index = {name: i for i, name in enumerate(header)}

    def get_value(row, column_name, default=""):
        i = index.get(column_name)
        if i is None or i >= len(row):
            return default
        value = row[i]
        return default if value is None else value

    imported = []
    for row in rows[1:]:
        entry_id = str(get_value(row, "ID", "")).strip()
        if not entry_id:
            continue

        imported.append({
            "id": entry_id,
            "date": get_value(row, "Date", ""),
            "timestamp": get_value(row, "Timestamp", ""),
            "event_name": get_value(row, "Event name", ""),
            "counted_by": get_value(row, "Counted by", ""),
            "cash_sum": get_value(row, "Cash sum", ""),
            "event_status": get_value(row, "Event status", ""),
            "comment": get_value(row, "Comment", ""),
            "denom_100": get_value(row, "100 €", ""),
            "denom_50": get_value(row, "50 €", ""),
            "denom_20": get_value(row, "20 €", ""),
            "denom_10": get_value(row, "10 €", ""),
            "denom_5": get_value(row, "5 €", ""),
            "denom_2": get_value(row, "2 €", ""),
            "denom_1": get_value(row, "1 €", ""),
            "denom_050": get_value(row, "0.50 €", ""),
            "denom_020": get_value(row, "0.20 €", ""),
            "denom_010": get_value(row, "0.10 €", ""),
        })

    return imported