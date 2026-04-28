from openpyxl import load_workbook

from core.export_utils import (
    export_entries_to_excel,
    export_entries_to_text,
    import_entries_from_excel,
)
from core.storage import ensure_db_file, insert_entry


def test_export_entries_to_excel_creates_expected_sheet(temp_paths, sample_entry):
    db_path = temp_paths["db_path"]
    excel_path = temp_paths["excel_path"]

    ensure_db_file(db_path)
    insert_entry(db_path, sample_entry)

    export_entries_to_excel(db_path, excel_path)

    assert excel_path.exists()

    workbook = load_workbook(excel_path, data_only=True)
    sheet = workbook.active

    assert sheet.title == "Kassensturz"
    assert sheet.max_row == 2
    assert sheet["A2"].value == "test-id-001"
    assert sheet["D2"].value == "Barabend"


def test_export_entries_to_text_creates_ascii_table(temp_paths, sample_entry):
    db_path = temp_paths["db_path"]
    text_path = temp_paths["text_path"]

    ensure_db_file(db_path)
    insert_entry(db_path, sample_entry)

    export_entries_to_text(db_path, text_path)

    assert text_path.exists()
    content = text_path.read_text(encoding="utf-8")

    assert "Barabend" in content
    assert "Jan" in content
    assert "test-id-001"[:12] in content
    assert "|" in content


def test_import_entries_from_excel_current_format(temp_paths, sample_entry):
    db_path = temp_paths["db_path"]
    excel_path = temp_paths["excel_path"]

    ensure_db_file(db_path)
    insert_entry(db_path, sample_entry)
    export_entries_to_excel(db_path, excel_path)

    imported = import_entries_from_excel(excel_path)

    assert len(imported) == 1
    assert imported[0]["id"] == "test-id-001"
    assert imported[0]["event_name"] == "Barabend"


def test_import_entries_from_legacy_excel_without_id(tmp_path):
    from openpyxl import Workbook

    legacy_file = tmp_path / "legacy.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.append([
        "Date",
        "Timestamp",
        "Event name",
        "Counted by",
        "Cash sum",
        "Event status",
        "Comment",
    ])
    sheet.append([
        "2026-04-28",
        "2026-04-28 10:00:00",
        "Barabend",
        "Jan",
        83.4,
        "opening",
        "legacy",
    ])
    workbook.save(legacy_file)

    imported = import_entries_from_excel(legacy_file)

    assert len(imported) == 1
    assert imported[0]["id"]
    assert imported[0]["event_name"] == "Barabend"
    assert imported[0]["counted_by"] == "Jan"
    assert imported[0]["comment"] == "legacy"