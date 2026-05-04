from datetime import date, datetime, time
from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.utils.datetime import to_excel

from core import export_utils
from core.storage_counts import create_cash_count
from core.storage_movements import create_cash_movement

FIXTURES_DIR = Path(__file__).parent / "fixtures"


def test_export_and_import_roundtrip(
    seeded_db,
    excel_path,
    text_path,
    bar_account_id,
    runner_account_id,
):
    create_cash_count(
        db_path=seeded_db,
        cash_account_id=bar_account_id,
        counted_by="Jan",
        total_cents=12345,
        count_type="opening",
        context_label="Friday Bar",
        denominations={"roll_2": 1},
    )

    create_cash_movement(
        db_path=seeded_db,
        from_account_id=bar_account_id,
        to_account_id=runner_account_id,
        amount_cents=5000,
        context_label="Friday Bar",
        denominations={"denom_20": 1, "denom_10": 1},
    )

    export_utils.export_all(
        db_path=seeded_db,
        excel_path=excel_path,
        text_path=text_path,
    )

    assert excel_path.exists()
    assert text_path.exists()
    text_export = text_path.read_text(encoding="utf-8")

    assert "=== CASH ACCOUNTS ===" in text_export
    assert "=== CASH BALANCES ===" in text_export
    assert "=== CASH MOVEMENTS ===" in text_export
    assert "=== CASH COUNTS ===" in text_export
    assert "amount_eur=50.00" in text_export
    assert "total_eur=123.45" in text_export
    assert "denoms=denom_20=1,denom_10=1" in text_export
    assert "denoms=roll_2=1" in text_export

    imported = export_utils.import_all_from_excel(excel_path)

    assert "cash_accounts" in imported
    assert "cash_contexts" in imported
    assert "cash_counts" in imported
    assert "cash_movements" in imported

    assert len(imported["cash_counts"]) == 1
    assert len(imported["cash_movements"]) == 1

    count_row = imported["cash_counts"][0]
    movement_row = imported["cash_movements"][0]

    assert count_row["total_cents"] == 12345
    assert count_row["roll_2"] == 1

    assert movement_row["amount_cents"] == 5000
    assert movement_row["denom_20"] == 1
    assert movement_row["denom_10"] == 1


def test_import_legacy_cash_count_workbook(tmp_path):
    excel_path = tmp_path / "legacy.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Production"
    worksheet.append(
        [
            "Date",
            "Timestamp",
            "Event name",
            "Counted by",
            "Cash sum",
            "Event status",
            "Comment",
        ]
    )
    worksheet.append(
        [
            date(2026, 4, 30),
            time(23, 45),
            "Friday Bar",
            "Jan",
            "1.234,56 €",
            "Closed",
            "legacy closing count",
        ]
    )
    workbook.save(excel_path)

    imported = export_utils.import_all_from_excel(excel_path)

    assert imported["source_format"] == "legacy_cash_counts"
    assert len(imported["cash_contexts"]) == 1
    assert imported["cash_contexts"][0]["label"] == "Friday Bar"
    assert len(imported["cash_counts"]) == 1

    count = imported["cash_counts"][0]
    assert count["cash_account_id"] == "acc_bar_cash_box"
    assert count["cash_account_name"] == "Bar Cash Box"
    assert count["counted_at"] == "2026-04-30T23:45:00"
    assert count["counted_by"] == "Jan"
    assert count["total_cents"] == 123456
    assert count["count_type"] == "closing"
    assert count["note"] == "legacy closing count"


def test_import_anonymized_legacy_cash_count_workbook_fixture():
    imported = export_utils.import_all_from_excel(
        FIXTURES_DIR / "legacy_cash_counts_anonymized.xlsx"
    )

    assert imported["source_format"] == "legacy_cash_counts"
    assert len(imported["cash_contexts"]) == 2
    assert len(imported["cash_counts"]) == 3

    contexts = {context["label"]: context for context in imported["cash_contexts"]}
    assert contexts["Event Alpha"]["last_used_at"] == "2026-05-02T18:30:15"
    assert contexts["Event Beta"]["last_used_at"] == "2026-05-01T01:15:00"

    first_count, second_count, third_count = imported["cash_counts"]

    assert first_count["counted_at"] == "2026-04-30T23:45:00"
    assert first_count["counted_by"] == "Person A"
    assert first_count["context_label"] == "Event Alpha"
    assert first_count["total_cents"] == 123456
    assert first_count["count_type"] == "closing"

    assert second_count["counted_at"] == "2026-05-01T01:15:00"
    assert second_count["counted_by"] == "Person B"
    assert second_count["context_label"] == "Event Beta"
    assert second_count["total_cents"] == 98765
    assert second_count["count_type"] == "opening"

    assert third_count["counted_at"] == "2026-05-02T18:30:15"
    assert third_count["counted_by"] == "Person C"
    assert third_count["context_label"] == "Event Alpha"
    assert third_count["total_cents"] == 111110
    assert third_count["count_type"] == "spot_check"


def test_import_missing_excel_returns_empty_missing_source(tmp_path):
    imported = export_utils.import_all_from_excel(tmp_path / "does-not-exist.xlsx")

    assert imported == {
        "cash_accounts": [],
        "cash_contexts": [],
        "cash_movements": [],
        "cash_counts": [],
        "source_format": "missing",
    }


def test_import_modern_workbook_skips_blank_rows_and_missing_sheets(tmp_path):
    excel_path = tmp_path / "modern-accounts-only.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = export_utils.SHEET_CASH_ACCOUNTS
    worksheet.append(
        [
            "id",
            "name",
            "account_type",
            "current_balance_cents",
            "is_active",
            "sort_order",
            "created_at",
        ]
    )
    worksheet.append([None, None, None, None, None, None, None])
    worksheet.append(
        [
            "acc_edge",
            "Edge Cash Box",
            "cash",
            12345,
            1,
            99,
            "2026-05-04T12:00:00",
        ]
    )
    workbook.save(excel_path)

    imported = export_utils.import_all_from_excel(excel_path)

    assert imported["source_format"] == "kassensturz"
    assert imported["cash_accounts"] == [
        {
            "id": "acc_edge",
            "name": "Edge Cash Box",
            "account_type": "cash",
            "current_balance_cents": 12345,
            "is_active": 1,
            "sort_order": 99,
            "created_at": "2026-05-04T12:00:00",
        }
    ]
    assert imported["cash_contexts"] == []
    assert imported["cash_movements"] == []
    assert imported["cash_counts"] == []


def test_import_modern_count_sheet_missing_columns_fills_none(tmp_path):
    excel_path = tmp_path / "modern-sparse-counts.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = export_utils.SHEET_CASH_COUNTS
    worksheet.append(["id", "total_cents"])
    worksheet.append(["count-sparse", 777])
    workbook.save(excel_path)

    imported = export_utils.import_all_from_excel(excel_path)

    assert imported["source_format"] == "kassensturz"
    assert len(imported["cash_counts"]) == 1

    count = imported["cash_counts"][0]
    assert count["id"] == "count-sparse"
    assert count["total_cents"] == 777
    assert count["context_id"] is None
    assert count["cash_account_id"] is None
    assert count["denom_100"] is None


def test_import_legacy_cash_count_workbook_ignores_blank_rows_and_numeric_dates(tmp_path):
    excel_path = tmp_path / "legacy-numeric-dates.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Production"
    worksheet.append(["Intro row before headers"])
    worksheet.append(
        [
            " DATE ",
            "Timestamp",
            "Event name",
            "Counted by",
            "Cash sum",
            "Event status",
            "Comment",
        ]
    )
    worksheet.append([None, None, None, None, None, None, None])
    worksheet.append(
        [
            to_excel(datetime(2026, 5, 2)),
            0.75,
            " Saturday Bar ",
            "",
            "12.345,67",
            "",
            None,
        ]
    )
    workbook.save(excel_path)

    imported = export_utils.import_all_from_excel(excel_path)

    assert imported["source_format"] == "legacy_cash_counts"
    assert len(imported["cash_contexts"]) == 1
    assert imported["cash_contexts"][0]["label"] == "Saturday Bar"
    assert len(imported["cash_counts"]) == 1

    count = imported["cash_counts"][0]
    assert count["counted_at"] == "2026-05-02T18:00:00"
    assert count["context_label"] == "Saturday Bar"
    assert count["count_type"] == "reconciliation"
    assert count["counted_by"] == ""
    assert count["total_cents"] == 1234567
    assert count["note"] == ""


def test_import_legacy_cash_count_workbook_uses_current_time_when_date_is_invalid(
    tmp_path,
    monkeypatch,
):
    excel_path = tmp_path / "legacy-invalid-date.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Production"
    worksheet.append(
        [
            "Date",
            "Timestamp",
            "Event name",
            "Counted by",
            "Cash sum",
            "Event status",
            "Comment",
        ]
    )
    worksheet.append(
        [
            "not a date",
            "not a time",
            "",
            "Jan",
            "1.00",
            "Abgleich",
            "",
        ]
    )
    workbook.save(excel_path)
    monkeypatch.setattr(export_utils, "now_iso", lambda: "2026-05-04T10:11:12")

    imported = export_utils.import_all_from_excel(excel_path)

    assert imported["cash_contexts"] == []
    assert imported["cash_counts"][0]["counted_at"] == "2026-05-04T10:11:12"
    assert imported["cash_counts"][0]["count_type"] == "reconciliation"


def test_import_legacy_cash_count_workbook_rejects_invalid_cash_sum(tmp_path):
    excel_path = tmp_path / "legacy-invalid-currency.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Production"
    worksheet.append(
        [
            "Date",
            "Timestamp",
            "Event name",
            "Counted by",
            "Cash sum",
            "Event status",
            "Comment",
        ]
    )
    worksheet.append(
        [
            date(2026, 5, 4),
            time(12, 0),
            "Monday Bar",
            "Jan",
            "not money",
            "closed",
            "",
        ]
    )
    workbook.save(excel_path)

    with pytest.raises(ValueError, match="Invalid legacy cash sum"):
        export_utils.import_all_from_excel(excel_path)
