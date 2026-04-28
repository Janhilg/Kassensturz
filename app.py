import hashlib
import mimetypes
import os
import shutil
import sqlite3
import sys
import tempfile
import threading
import uuid
import webbrowser
from datetime import datetime
from pathlib import Path
from urllib.parse import quote

import requests
from flask import Flask, flash, redirect, render_template, request, url_for
from openpyxl import Workbook, load_workbook

from config import Config

mimetypes.add_type("application/javascript", ".js")

print(f"[Kassensturz] MODE = {Config.MODE}")
print(f"[Kassensturz] FROZEN = {getattr(sys, 'frozen', False)}")


def resource_path(relative_path: str) -> str:
    if hasattr(sys, "_MEIPASS"):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.abspath("."), relative_path)


def portable_base_dir() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def is_debug_mode():
    return Config.MODE == "debug" and not Config.IS_FROZEN


template_dir = resource_path("templates")
static_dir = resource_path("static")

app = Flask(__name__, template_folder=template_dir, static_folder=static_dir)
app.secret_key = os.getenv("FLASK_SECRET_KEY", "change-this-secret-key")

EXCEL_HEADERS = [
    "ID",
    "Date",
    "Timestamp",
    "Event name",
    "Counted by",
    "Cash sum",
    "Event status",
    "Comment",
    "100 €",
    "50 €",
    "20 €",
    "10 €",
    "5 €",
    "2 €",
    "1 €",
    "0.50 €",
    "0.20 €",
    "0.10 €",
]

DB_COLUMNS = [
    "id",
    "date",
    "timestamp",
    "event_name",
    "counted_by",
    "cash_sum",
    "event_status",
    "comment",
    "denom_100",
    "denom_50",
    "denom_20",
    "denom_10",
    "denom_5",
    "denom_2",
    "denom_1",
    "denom_050",
    "denom_020",
    "denom_010",
]

BASE_DIR = portable_base_dir()

if is_debug_mode():
    LOCAL_DB_FILE = BASE_DIR / "data_debug" / "kassensturz.db"
    BACKUP_DIR = BASE_DIR / "data_debug" / "backups"
else:
    LOCAL_DB_FILE = BASE_DIR / "data" / "kassensturz.db"
    BACKUP_DIR = BASE_DIR / "data" / "backups"

NEXTCLOUD_BASE_URL = Config.NEXTCLOUD_BASE_URL.rstrip("/")
NEXTCLOUD_USERNAME = Config.NEXTCLOUD_USERNAME
NEXTCLOUD_APP_PASSWORD = Config.NEXTCLOUD_APP_PASSWORD
NEXTCLOUD_REMOTE_DIR = Config.NEXTCLOUD_REMOTE_DIR
NEXTCLOUD_REMOTE_FILE = Config.NEXTCLOUD_REMOTE_FILE


def get_verify_setting():
    if getattr(Config, "NEXTCLOUD_VERIFY", "true").lower() == "false":
        return False

    ca_cert_path = getattr(Config, "NEXTCLOUD_CA_CERT_PATH", "")
    if ca_cert_path:
        ca_path = Path(ca_cert_path)
        if not ca_path.is_absolute():
            ca_path = BASE_DIR / ca_path
        return str(ca_path)

    return True


def nextcloud_configured():
    return all([
        NEXTCLOUD_BASE_URL,
        NEXTCLOUD_USERNAME,
        NEXTCLOUD_APP_PASSWORD,
    ])


def build_webdav_url(path: str) -> str:
    encoded_path = "/".join(quote(part) for part in path.strip("/").split("/"))
    return (
        f"{NEXTCLOUD_BASE_URL}/remote.php/dav/files/"
        f"{quote(NEXTCLOUD_USERNAME)}/{encoded_path}"
    )


def ensure_nextcloud_folder():
    if not nextcloud_configured():
        return

    parts = [part for part in NEXTCLOUD_REMOTE_DIR.strip("/").split("/") if part]
    current_path = ""

    for part in parts:
        current_path = f"{current_path}/{part}" if current_path else part
        url = build_webdav_url(current_path)

        response = requests.request(
            "MKCOL",
            url,
            auth=(NEXTCLOUD_USERNAME, NEXTCLOUD_APP_PASSWORD),
            timeout=30,
            verify=get_verify_setting(),
        )

        if response.status_code not in (201, 405):
            raise RuntimeError(
                f"Failed to create Nextcloud folder '{current_path}': "
                f"{response.status_code} {response.text}"
            )


def get_connection(db_path: Path):
    db_path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    return conn


def ensure_local_db_file():
    with get_connection(LOCAL_DB_FILE) as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS entries (
                id TEXT PRIMARY KEY,
                date TEXT NOT NULL,
                timestamp TEXT NOT NULL,
                event_name TEXT NOT NULL,
                counted_by TEXT NOT NULL,
                cash_sum REAL NOT NULL,
                event_status TEXT NOT NULL,
                comment TEXT DEFAULT '',
                denom_100 INTEGER,
                denom_50 INTEGER,
                denom_20 INTEGER,
                denom_10 INTEGER,
                denom_5 INTEGER,
                denom_2 INTEGER,
                denom_1 INTEGER,
                denom_050 INTEGER,
                denom_020 INTEGER,
                denom_010 INTEGER
            )
        """)
        conn.commit()


def new_entry_id():
    return str(uuid.uuid4())


def parse_optional_int(raw_value):
    raw_value = str(raw_value).strip()
    if raw_value == "":
        return None
    return int(raw_value)


def get_denomination_values_from_form(form):
    return {
        "denom_100": parse_optional_int(form.get("denom_100", "")),
        "denom_50": parse_optional_int(form.get("denom_50", "")),
        "denom_20": parse_optional_int(form.get("denom_20", "")),
        "denom_10": parse_optional_int(form.get("denom_10", "")),
        "denom_5": parse_optional_int(form.get("denom_5", "")),
        "denom_2": parse_optional_int(form.get("denom_2", "")),
        "denom_1": parse_optional_int(form.get("denom_1", "")),
        "denom_050": parse_optional_int(form.get("denom_050", "")),
        "denom_020": parse_optional_int(form.get("denom_020", "")),
        "denom_010": parse_optional_int(form.get("denom_010", "")),
    }


def insert_entry(entry: dict):
    ensure_local_db_file()

    values = [entry.get(column) for column in DB_COLUMNS]
    placeholders = ", ".join("?" for _ in DB_COLUMNS)
    columns_sql = ", ".join(DB_COLUMNS)

    with get_connection(LOCAL_DB_FILE) as conn:
        conn.execute(
            f"INSERT INTO entries ({columns_sql}) VALUES ({placeholders})",
            values,
        )
        conn.commit()


def fetch_all_entries():
    ensure_local_db_file()

    with get_connection(LOCAL_DB_FILE) as conn:
        rows = conn.execute(
            "SELECT * FROM entries ORDER BY timestamp ASC, id ASC"
        ).fetchall()
        return [dict(row) for row in rows]


def merge_imported_entries_append_only(imported_entries):
    ensure_local_db_file()

    with get_connection(LOCAL_DB_FILE) as conn:
        existing_ids = {
            row["id"]
            for row in conn.execute("SELECT id FROM entries").fetchall()
        }

        for entry in imported_entries:
            entry_id = str(entry.get("id", "")).strip()
            if not entry_id or entry_id in existing_ids:
                continue

            values = [entry.get(column) for column in DB_COLUMNS]
            placeholders = ", ".join("?" for _ in DB_COLUMNS)
            columns_sql = ", ".join(DB_COLUMNS)

            conn.execute(
                f"INSERT INTO entries ({columns_sql}) VALUES ({placeholders})",
                values,
            )
            existing_ids.add(entry_id)

        conn.commit()


def create_backup(max_backups: int = 25):
    ensure_local_db_file()
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_file = BACKUP_DIR / f"kassensturz_backup_{timestamp}.db"
    shutil.copy2(LOCAL_DB_FILE, backup_file)

    backups = sorted(
        BACKUP_DIR.glob("kassensturz_backup_*.db"),
        key=lambda f: f.stat().st_mtime,
        reverse=True,
    )

    for old_file in backups[max_backups:]:
        try:
            old_file.unlink()
        except Exception:
            pass

    return backup_file


def excel_safe_value(value):
    return "" if value is None else value


def export_entries_to_excel(file_path: Path):
    entries = fetch_all_entries()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Kassensturz"
    sheet.append(EXCEL_HEADERS)

    for entry in entries:
        sheet.append([
            excel_safe_value(entry.get("id")),
            excel_safe_value(entry.get("date")),
            excel_safe_value(entry.get("timestamp")),
            excel_safe_value(entry.get("event_name")),
            excel_safe_value(entry.get("counted_by")),
            excel_safe_value(entry.get("cash_sum")),
            excel_safe_value(entry.get("event_status")),
            excel_safe_value(entry.get("comment")),
            excel_safe_value(entry.get("denom_100")),
            excel_safe_value(entry.get("denom_50")),
            excel_safe_value(entry.get("denom_20")),
            excel_safe_value(entry.get("denom_10")),
            excel_safe_value(entry.get("denom_5")),
            excel_safe_value(entry.get("denom_2")),
            excel_safe_value(entry.get("denom_1")),
            excel_safe_value(entry.get("denom_050")),
            excel_safe_value(entry.get("denom_020")),
            excel_safe_value(entry.get("denom_010")),
        ])

    file_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(file_path)


def legacy_row_id(values: list):
    normalized = "|".join("" if value is None else str(value) for value in values)
    return hashlib.sha256(normalized.encode("utf-8")).hexdigest()


def get_cell_value(row, index_map, column_name, default=""):
    idx = index_map.get(column_name)
    if idx is None or idx >= len(row):
        return default
    value = row[idx]
    return default if value is None else value


def import_entries_from_excel(file_path: Path):
    workbook = load_workbook(file_path, data_only=True)
    sheet = workbook.active

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    header = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    index_map = {name: idx for idx, name in enumerate(header)}

    imported_entries = []

    for row in rows[1:]:
        row = list(row)

        # Current format with ID
        if "ID" in index_map:
            entry_id = str(get_cell_value(row, index_map, "ID", "")).strip()
            if not entry_id:
                continue

            imported_entries.append({
                "id": entry_id,
                "date": get_cell_value(row, index_map, "Date", ""),
                "timestamp": get_cell_value(row, index_map, "Timestamp", ""),
                "event_name": get_cell_value(row, index_map, "Event name", ""),
                "counted_by": get_cell_value(row, index_map, "Counted by", ""),
                "cash_sum": get_cell_value(row, index_map, "Cash sum", ""),
                "event_status": get_cell_value(row, index_map, "Event status", ""),
                "comment": get_cell_value(row, index_map, "Comment", ""),
                "denom_100": get_cell_value(row, index_map, "100 €", None),
                "denom_50": get_cell_value(row, index_map, "50 €", None),
                "denom_20": get_cell_value(row, index_map, "20 €", None),
                "denom_10": get_cell_value(row, index_map, "10 €", None),
                "denom_5": get_cell_value(row, index_map, "5 €", None),
                "denom_2": get_cell_value(row, index_map, "2 €", None),
                "denom_1": get_cell_value(row, index_map, "1 €", None),
                "denom_050": get_cell_value(row, index_map, "0.50 €", None),
                "denom_020": get_cell_value(row, index_map, "0.20 €", None),
                "denom_010": get_cell_value(row, index_map, "0.10 €", None),
            })
            continue

        # Older format without ID but with "Counted by"
        if all(col in index_map for col in [
            "Date", "Timestamp", "Event name", "Counted by", "Cash sum", "Event status", "Comment"
        ]):
            legacy_values = [
                get_cell_value(row, index_map, "Date", ""),
                get_cell_value(row, index_map, "Timestamp", ""),
                get_cell_value(row, index_map, "Event name", ""),
                get_cell_value(row, index_map, "Counted by", ""),
                get_cell_value(row, index_map, "Cash sum", ""),
                get_cell_value(row, index_map, "Event status", ""),
                get_cell_value(row, index_map, "Comment", ""),
                get_cell_value(row, index_map, "100 €", ""),
                get_cell_value(row, index_map, "50 €", ""),
                get_cell_value(row, index_map, "20 €", ""),
                get_cell_value(row, index_map, "10 €", ""),
                get_cell_value(row, index_map, "5 €", ""),
                get_cell_value(row, index_map, "2 €", ""),
                get_cell_value(row, index_map, "1 €", ""),
                get_cell_value(row, index_map, "0.50 €", ""),
                get_cell_value(row, index_map, "0.20 €", ""),
                get_cell_value(row, index_map, "0.10 €", ""),
            ]

            imported_entries.append({
                "id": legacy_row_id(legacy_values),
                "date": legacy_values[0],
                "timestamp": legacy_values[1],
                "event_name": legacy_values[2],
                "counted_by": legacy_values[3],
                "cash_sum": legacy_values[4],
                "event_status": legacy_values[5],
                "comment": legacy_values[6],
                "denom_100": legacy_values[7] or None,
                "denom_50": legacy_values[8] or None,
                "denom_20": legacy_values[9] or None,
                "denom_10": legacy_values[10] or None,
                "denom_5": legacy_values[11] or None,
                "denom_2": legacy_values[12] or None,
                "denom_1": legacy_values[13] or None,
                "denom_050": legacy_values[14] or None,
                "denom_020": legacy_values[15] or None,
                "denom_010": legacy_values[16] or None,
            })
            continue

        # Very old format without "Counted by"
        if all(col in index_map for col in [
            "Date", "Timestamp", "Event name", "Cash sum", "Event status", "Comment"
        ]):
            legacy_values = [
                get_cell_value(row, index_map, "Date", ""),
                get_cell_value(row, index_map, "Timestamp", ""),
                get_cell_value(row, index_map, "Event name", ""),
                "",
                get_cell_value(row, index_map, "Cash sum", ""),
                get_cell_value(row, index_map, "Event status", ""),
                get_cell_value(row, index_map, "Comment", ""),
                get_cell_value(row, index_map, "100 €", ""),
                get_cell_value(row, index_map, "50 €", ""),
                get_cell_value(row, index_map, "20 €", ""),
                get_cell_value(row, index_map, "10 €", ""),
                get_cell_value(row, index_map, "5 €", ""),
                get_cell_value(row, index_map, "2 €", ""),
                get_cell_value(row, index_map, "1 €", ""),
                get_cell_value(row, index_map, "0.50 €", ""),
                get_cell_value(row, index_map, "0.20 €", ""),
                get_cell_value(row, index_map, "0.10 €", ""),
            ]

            imported_entries.append({
                "id": legacy_row_id(legacy_values),
                "date": legacy_values[0],
                "timestamp": legacy_values[1],
                "event_name": legacy_values[2],
                "counted_by": legacy_values[3],
                "cash_sum": legacy_values[4],
                "event_status": legacy_values[5],
                "comment": legacy_values[6],
                "denom_100": legacy_values[7] or None,
                "denom_50": legacy_values[8] or None,
                "denom_20": legacy_values[9] or None,
                "denom_10": legacy_values[10] or None,
                "denom_5": legacy_values[11] or None,
                "denom_2": legacy_values[12] or None,
                "denom_1": legacy_values[13] or None,
                "denom_050": legacy_values[14] or None,
                "denom_020": legacy_values[15] or None,
                "denom_010": legacy_values[16] or None,
            })

    return imported_entries


def download_remote_to_temp(temp_path: Path):
    if not nextcloud_configured():
        return False

    remote_path = f"{NEXTCLOUD_REMOTE_DIR}/{NEXTCLOUD_REMOTE_FILE}"
    url = build_webdav_url(remote_path)

    response = requests.get(
        url,
        auth=(NEXTCLOUD_USERNAME, NEXTCLOUD_APP_PASSWORD),
        timeout=60,
        verify=get_verify_setting(),
    )

    if response.status_code == 200:
        temp_path.parent.mkdir(parents=True, exist_ok=True)
        with temp_path.open("wb") as file_handle:
            file_handle.write(response.content)
        return True

    if response.status_code == 404:
        return False

    raise RuntimeError(
        f"Failed to download Excel file from Nextcloud: "
        f"{response.status_code} {response.text}"
    )


def upload_excel_file_to_nextcloud(file_path: Path):
    if not nextcloud_configured():
        return

    ensure_nextcloud_folder()

    remote_path = f"{NEXTCLOUD_REMOTE_DIR}/{NEXTCLOUD_REMOTE_FILE}"
    url = build_webdav_url(remote_path)

    print(f"[Kassensturz] Uploading to: {remote_path}")
    print(f"[Kassensturz] WebDAV URL: {url}")

    with file_path.open("rb") as file_handle:
        response = requests.put(
            url,
            data=file_handle,
            auth=(NEXTCLOUD_USERNAME, NEXTCLOUD_APP_PASSWORD),
            headers={
                "Content-Type": (
                    "application/vnd.openxmlformats-officedocument."
                    "spreadsheetml.sheet"
                )
            },
            timeout=60,
            verify=get_verify_setting(),
        )

    print(f"[Kassensturz] Upload response: {response.status_code}")
    print(f"[Kassensturz] Upload response body: {response.text[:500]}")

    if response.status_code not in (200, 201, 204):
        raise RuntimeError(
            f"Failed to upload Excel file to Nextcloud: "
            f"{response.status_code} {response.text}"
        )


def append_and_sync(entry: dict):
    ensure_local_db_file()
    insert_entry(entry)

    if nextcloud_configured():
        flash("upload_success", "success")

    if not nextcloud_configured():
        return

    with tempfile.TemporaryDirectory() as tmp_dir_name:
        tmp_dir = Path(tmp_dir_name)
        remote_excel_file = tmp_dir / "remote.xlsx"
        merged_export_file = tmp_dir / "kassensturz_data.xlsx"

        remote_exists = False

        if not is_debug_mode():
            remote_exists = download_remote_to_temp(remote_excel_file)

        if remote_exists:
            imported_entries = import_entries_from_excel(remote_excel_file)
            merge_imported_entries_append_only(imported_entries)

        create_backup()
        export_entries_to_excel(merged_export_file)
        upload_excel_file_to_nextcloud(merged_export_file)


@app.route("/", methods=["GET", "POST"])
def home():
    if request.method == "POST":
        submitted_text = request.form.get("text_input", "").strip()
        submitted_counted_by = request.form.get("counted_by_input", "").strip()
        submitted_number = request.form.get("number_input", "").strip()
        submitted_event_state = request.form.get("event_state", "").strip()
        submitted_comment = request.form.get("comment_input", "").strip()

        try:
            denominations = get_denomination_values_from_form(request.form)
        except Exception:
            flash("Invalid denomination input.", "error")
            return redirect(url_for("home"))

        if submitted_text and submitted_counted_by and submitted_number and submitted_event_state:
            try:
                now = datetime.now()
                current_date = now.strftime("%Y-%m-%d")
                current_timestamp = now.strftime("%Y-%m-%d %H:%M:%S")

                entry = {
                    "id": new_entry_id(),
                    "date": current_date,
                    "timestamp": current_timestamp,
                    "event_name": submitted_text,
                    "counted_by": submitted_counted_by,
                    "cash_sum": float(submitted_number),
                    "event_status": submitted_event_state,
                    "comment": submitted_comment,
                    "denom_100": denominations["denom_100"],
                    "denom_50": denominations["denom_50"],
                    "denom_20": denominations["denom_20"],
                    "denom_10": denominations["denom_10"],
                    "denom_5": denominations["denom_5"],
                    "denom_2": denominations["denom_2"],
                    "denom_1": denominations["denom_1"],
                    "denom_050": denominations["denom_050"],
                    "denom_020": denominations["denom_020"],
                    "denom_010": denominations["denom_010"],
                }

                append_and_sync(entry)

                flash(
                    {
                        "date": current_date,
                        "timestamp": current_timestamp,
                        "text": submitted_text,
                        "counted_by": submitted_counted_by,
                        "number": submitted_number,
                        "event_state": submitted_event_state,
                        "comment": submitted_comment,
                        "denominations": denominations,
                    },
                    "submitted",
                )

                return redirect(url_for("home"))

            except Exception as exc:
                flash(str(exc), "error")
                return redirect(url_for("home"))

    return render_template("index.html", app_mode=Config.MODE)


def open_browser():
    webbrowser.open("http://127.0.0.1:5000")


if __name__ == "__main__":
    ensure_local_db_file()
    threading.Timer(1.0, open_browser).start()
    app.run(host="127.0.0.1", port=5000, debug=is_debug_mode())