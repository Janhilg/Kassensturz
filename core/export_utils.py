import logging
import uuid
from datetime import date, datetime, time
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation
from pathlib import Path

from openpyxl import Workbook, load_workbook

from core import storage

logger = logging.getLogger(__name__)


SHEET_CASH_ACCOUNTS = "CashAccounts"
SHEET_CASH_CONTEXTS = "CashContexts"
SHEET_CASH_MOVEMENTS = "CashMovements"
SHEET_CASH_COUNTS = "CashCounts"
SHEET_CASH_BALANCES = "CashBalances"

LEGACY_CASH_COUNT_HEADERS = {
    "date": "date",
    "timestamp": "timestamp",
    "event name": "context_label",
    "counted by": "counted_by",
    "cash sum": "cash_sum",
    "event status": "count_type",
    "comment": "note",
}
LEGACY_REQUIRED_HEADERS = set(LEGACY_CASH_COUNT_HEADERS)
LEGACY_IMPORT_NAMESPACE = uuid.UUID("f4e0f289-76f0-4f03-995d-9a0a92f61735")


def _ensure_parent_dir(path: Path):
    path.parent.mkdir(parents=True, exist_ok=True)


def _write_sheet(ws, columns: list[str], rows: list[dict]):
    ws.append(columns)
    for row in rows:
        ws.append([row.get(column) for column in columns])


def _read_sheet_rows(workbook, sheet_name: str) -> list[dict]:
    if sheet_name not in workbook.sheetnames:
        return []

    ws = workbook[sheet_name]
    values = list(ws.values)
    if not values:
        return []

    header = list(values[0])
    rows = []

    for raw_row in values[1:]:
        if raw_row is None:
            continue

        row = dict(zip(header, raw_row, strict=False))

        if all(value is None for value in row.values()):
            continue

        cleaned = {}
        for key, value in row.items():
            if value is None:
                cleaned[key] = None
            else:
                cleaned[key] = value

        rows.append(cleaned)

    return rows


def _safe_str(value) -> str:
    if value is None:
        return ""
    return str(value)


def _normalize_legacy_header(value) -> str:
    return " ".join(_safe_str(value).strip().split()).casefold()


def _legacy_uuid(*parts) -> str:
    normalized = "|".join(_safe_str(part).strip() for part in parts)
    return f"legacy-{uuid.uuid5(LEGACY_IMPORT_NAMESPACE, normalized)}"


def _parse_legacy_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = _safe_str(value).strip()
    if not text:
        return None

    for date_format in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, date_format).date()
        except ValueError:
            continue

    try:
        return datetime.fromisoformat(text).date()
    except ValueError:
        return None


def _parse_legacy_time(value):
    if isinstance(value, datetime):
        return value.time().replace(microsecond=0)
    if isinstance(value, time):
        return value.replace(microsecond=0)

    text = _safe_str(value).strip()
    if not text:
        return None

    for time_format in ("%H:%M:%S", "%H:%M"):
        try:
            return datetime.strptime(text, time_format).time()
        except ValueError:
            continue

    try:
        return datetime.fromisoformat(text).time().replace(microsecond=0)
    except ValueError:
        return None


def _legacy_counted_at(raw_date, raw_timestamp) -> str:
    if isinstance(raw_timestamp, datetime):
        return raw_timestamp.isoformat(timespec="seconds")

    parsed_date = _parse_legacy_date(raw_date)
    parsed_time = _parse_legacy_time(raw_timestamp)

    if parsed_date and parsed_time:
        return datetime.combine(parsed_date, parsed_time).isoformat(timespec="seconds")

    if parsed_date:
        return datetime.combine(parsed_date, time()).isoformat(timespec="seconds")

    return storage.now_iso()


def _parse_legacy_cash_sum(value) -> int:
    if isinstance(value, int | float | Decimal):
        amount = Decimal(str(value))
    else:
        text = _safe_str(value).strip()
        if not text:
            raise ValueError("Legacy cash sum is empty")

        text = text.replace("EUR", "").replace("€", "").replace(" ", "")
        if "," in text and "." in text:
            if text.rfind(",") > text.rfind("."):
                text = text.replace(".", "").replace(",", ".")
            else:
                text = text.replace(",", "")
        elif "," in text:
            text = text.replace(",", ".")

        try:
            amount = Decimal(text)
        except InvalidOperation as exc:
            raise ValueError(f"Invalid legacy cash sum: {value}") from exc

    cents = (amount * 100).quantize(Decimal("1"), rounding=ROUND_HALF_UP)
    if cents < 0:
        raise ValueError(f"Legacy cash sum cannot be negative: {value}")

    return int(cents)


def _normalize_legacy_count_type(value) -> str:
    raw_value = storage.normalize_optional_text(value)
    if not raw_value:
        return "reconciliation"

    key = raw_value.casefold().replace("-", "_").replace(" ", "_")
    count_type_map = {
        "open": "opening",
        "opening": "opening",
        "opened": "opening",
        "start": "opening",
        "closed": "closing",
        "close": "closing",
        "closing": "closing",
        "end": "closing",
        "spot": "spot_check",
        "spot_check": "spot_check",
        "spotcheck": "spot_check",
        "reconcile": "reconciliation",
        "reconciliation": "reconciliation",
    }
    return count_type_map.get(key, key)


def _legacy_row_values(raw_row, header_indexes: dict[int, str]) -> dict:
    row = {}
    for index, field_name in header_indexes.items():
        row[field_name] = raw_row[index] if index < len(raw_row) else None
    return row


def _import_legacy_cash_count_workbook(workbook) -> dict | None:
    for worksheet in workbook.worksheets:
        values = list(worksheet.values)
        if not values:
            continue

        for header_row_index, raw_header in enumerate(values[:10]):
            normalized_headers = [_normalize_legacy_header(value) for value in raw_header]
            if not LEGACY_REQUIRED_HEADERS.issubset(set(normalized_headers)):
                continue

            header_indexes = {
                index: LEGACY_CASH_COUNT_HEADERS[header]
                for index, header in enumerate(normalized_headers)
                if header in LEGACY_CASH_COUNT_HEADERS
            }
            imported_contexts_by_id = {}
            imported_counts = []

            for raw_row in values[header_row_index + 1 :]:
                legacy_row = _legacy_row_values(raw_row, header_indexes)
                if all(value in (None, "") for value in legacy_row.values()):
                    continue

                counted_at = _legacy_counted_at(
                    legacy_row.get("date"),
                    legacy_row.get("timestamp"),
                )
                context_label = storage.normalize_context_label(legacy_row.get("context_label"))
                context_id = None
                if context_label:
                    context_id = _legacy_uuid("context", context_label)
                    existing_context = imported_contexts_by_id.get(context_id)
                    if existing_context:
                        existing_context["last_used_at"] = max(
                            existing_context["last_used_at"],
                            counted_at,
                        )
                    else:
                        imported_contexts_by_id[context_id] = {
                            "id": context_id,
                            "label": context_label,
                            "created_at": counted_at,
                            "last_used_at": counted_at,
                            "is_active": 1,
                        }

                count_id = _legacy_uuid(
                    "count",
                    counted_at,
                    context_label,
                    legacy_row.get("counted_by"),
                    legacy_row.get("cash_sum"),
                    legacy_row.get("count_type"),
                    legacy_row.get("note"),
                )
                count_record = {
                    "id": count_id,
                    "context_id": context_id,
                    "context_label": context_label,
                    "cash_account_id": "acc_bar_cash_box",
                    "cash_account_name": "Bar Cash Box",
                    "counted_at": counted_at,
                    "count_type": _normalize_legacy_count_type(legacy_row.get("count_type")),
                    "counted_by": storage.normalize_optional_text(legacy_row.get("counted_by")),
                    "total_cents": _parse_legacy_cash_sum(legacy_row.get("cash_sum")),
                    "note": storage.normalize_optional_text(legacy_row.get("note")),
                }

                for field in storage.DENOM_FIELDS:
                    count_record[field] = None

                imported_counts.append(count_record)

            return {
                "cash_accounts": [],
                "cash_contexts": list(imported_contexts_by_id.values()),
                "cash_movements": [],
                "cash_counts": imported_counts,
                "source_format": "legacy_cash_counts",
            }

    return None


def _format_cents(cents) -> str:
    if cents is None:
        return ""
    return f"{int(cents) / 100:.2f}"


def export_excel(db_path: Path, excel_path: Path):
    storage.ensure_db_file(db_path)
    _ensure_parent_dir(excel_path)

    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    accounts = storage.fetch_all_cash_accounts(db_path, active_only=False)
    contexts = storage.fetch_recent_cash_contexts(db_path, limit=10000)
    movements = storage.fetch_all_cash_movements(db_path)
    counts = storage.fetch_all_cash_counts(db_path)
    balances = storage.fetch_cash_account_balances(db_path)

    ws_accounts = wb.create_sheet(SHEET_CASH_ACCOUNTS)
    _write_sheet(ws_accounts, storage.CASH_ACCOUNT_COLUMNS, accounts)

    ws_contexts = wb.create_sheet(SHEET_CASH_CONTEXTS)
    _write_sheet(ws_contexts, storage.CASH_CONTEXT_COLUMNS, contexts)

    movement_export_columns = [
        *storage.CASH_MOVEMENT_COLUMNS,
        "from_account_name",
        "to_account_name",
        "amount_eur",
    ]
    movement_export_rows = []
    for row in movements:
        movement_export_rows.append(
            {
                **row,
                "amount_eur": storage.cents_to_eur(int(row["amount_cents"])),
            }
        )

    ws_movements = wb.create_sheet(SHEET_CASH_MOVEMENTS)
    _write_sheet(ws_movements, movement_export_columns, movement_export_rows)

    count_export_columns = [
        *storage.CASH_COUNT_COLUMNS,
        "cash_account_name",
        "total_eur",
    ]
    count_export_rows = []
    for row in counts:
        count_export_rows.append(
            {
                **row,
                "total_eur": storage.cents_to_eur(int(row["total_cents"])),
            }
        )

    ws_counts = wb.create_sheet(SHEET_CASH_COUNTS)
    _write_sheet(ws_counts, count_export_columns, count_export_rows)

    balance_columns = [
        "id",
        "name",
        "account_type",
        "balance_cents",
        "balance_eur",
        "is_active",
        "sort_order",
    ]
    ws_balances = wb.create_sheet(SHEET_CASH_BALANCES)
    _write_sheet(ws_balances, balance_columns, balances)

    wb.save(excel_path)

    logger.info(
        "Excel export completed | path=%s accounts=%s contexts=%s movements=%s counts=%s",
        excel_path,
        len(accounts),
        len(contexts),
        len(movements),
        len(counts),
    )


def export_text(db_path: Path, text_path: Path):
    storage.ensure_db_file(db_path)
    _ensure_parent_dir(text_path)

    accounts = storage.fetch_all_cash_accounts(db_path, active_only=False)
    balances = storage.fetch_cash_account_balances(db_path)
    movements = storage.fetch_all_cash_movements(db_path)
    counts = storage.fetch_all_cash_counts(db_path)

    lines = []

    lines.append("=== CASH ACCOUNTS ===")
    for row in accounts:
        lines.append(
            " | ".join(
                [
                    _safe_str(row["name"]),
                    _safe_str(row["account_type"]),
                    f"active={row['is_active']}",
                    f"sort={row['sort_order']}",
                    _safe_str(row["id"]),
                ]
            )
        )

    lines.append("")
    lines.append("=== CASH BALANCES ===")
    for row in balances:
        lines.append(
            " | ".join(
                [
                    _safe_str(row["name"]),
                    _safe_str(row["account_type"]),
                    f"balance_eur={row['balance_eur']:.2f}",
                    f"balance_cents={row['balance_cents']}",
                ]
            )
        )

    lines.append("")
    lines.append("=== CASH MOVEMENTS ===")
    for row in movements:
        denom_summary = []
        for field in storage.DENOM_FIELDS:
            value = row.get(field)
            if value not in (None, "", 0):
                denom_summary.append(f"{field}={value}")

        lines.append(
            " | ".join(
                [
                    _safe_str(row["effective_at"]),
                    f"amount_eur={_format_cents(row['amount_cents'])}",
                    f"from={_safe_str(row.get('from_account_name'))}",
                    f"to={_safe_str(row.get('to_account_name'))}",
                    f"context={_safe_str(row.get('context_label'))}",
                    f"actor={_safe_str(row.get('actor'))}",
                    f"reference={_safe_str(row.get('reference'))}",
                    f"note={_safe_str(row.get('note'))}",
                    f"denoms={','.join(denom_summary)}",
                    f"id={_safe_str(row.get('id'))}",
                ]
            )
        )

    lines.append("")
    lines.append("=== CASH COUNTS ===")
    for row in counts:
        denom_summary = []
        for field in storage.DENOM_FIELDS:
            value = row.get(field)
            if value not in (None, "", 0):
                denom_summary.append(f"{field}={value}")

        lines.append(
            " | ".join(
                [
                    _safe_str(row["counted_at"]),
                    _safe_str(row["count_type"]),
                    f"total_eur={_format_cents(row['total_cents'])}",
                    f"account={_safe_str(row.get('cash_account_name'))}",
                    f"counted_by={_safe_str(row.get('counted_by'))}",
                    f"context={_safe_str(row.get('context_label'))}",
                    f"note={_safe_str(row.get('note'))}",
                    f"denoms={','.join(denom_summary)}",
                    f"id={_safe_str(row.get('id'))}",
                ]
            )
        )

    text_path.write_text("\n".join(lines) + "\n", encoding="utf-8")

    logger.info(
        "Text export completed | path=%s balances=%s movements=%s counts=%s",
        text_path,
        len(balances),
        len(movements),
        len(counts),
    )


def export_all(
    db_path: Path,
    excel_path: Path,
    text_path: Path,
):
    export_excel(db_path, excel_path)
    export_text(db_path, text_path)


def import_all_from_excel(excel_path: Path) -> dict:
    if not excel_path.exists():
        return {
            "cash_accounts": [],
            "cash_contexts": [],
            "cash_movements": [],
            "cash_counts": [],
            "source_format": "missing",
        }

    wb = load_workbook(excel_path, data_only=True)

    account_rows = _read_sheet_rows(wb, SHEET_CASH_ACCOUNTS)
    context_rows = _read_sheet_rows(wb, SHEET_CASH_CONTEXTS)
    movement_rows = _read_sheet_rows(wb, SHEET_CASH_MOVEMENTS)
    count_rows = _read_sheet_rows(wb, SHEET_CASH_COUNTS)

    if not any([account_rows, context_rows, movement_rows, count_rows]):
        legacy_import = _import_legacy_cash_count_workbook(wb)
        if legacy_import is not None:
            logger.info(
                "Legacy cash count Excel import completed | path=%s counts=%s contexts=%s",
                excel_path,
                len(legacy_import["cash_counts"]),
                len(legacy_import["cash_contexts"]),
            )
            return legacy_import

    imported_accounts = []
    for row in account_rows:
        imported_accounts.append(
            {column: row.get(column) for column in storage.CASH_ACCOUNT_COLUMNS}
        )

    imported_contexts = []
    for row in context_rows:
        imported_contexts.append(
            {column: row.get(column) for column in storage.CASH_CONTEXT_COLUMNS}
        )

    imported_movements = []
    for row in movement_rows:
        imported_movements.append(
            {column: row.get(column) for column in storage.CASH_MOVEMENT_COLUMNS}
        )

    imported_counts = []
    for row in count_rows:
        imported_counts.append({column: row.get(column) for column in storage.CASH_COUNT_COLUMNS})

    logger.info(
        "Excel import completed | path=%s accounts=%s contexts=%s movements=%s counts=%s",
        excel_path,
        len(imported_accounts),
        len(imported_contexts),
        len(imported_movements),
        len(imported_counts),
    )

    return {
        "cash_accounts": imported_accounts,
        "cash_contexts": imported_contexts,
        "cash_movements": imported_movements,
        "cash_counts": imported_counts,
        "source_format": "kassensturz",
    }


def __getattr__(name: str):
    if name == "CashExportService":
        from core.cash_export_service import CashExportService

        return CashExportService

    raise AttributeError(name)


__all__ = [
    "export_all",
    "export_excel",
    "export_text",
    "import_all_from_excel",
]
